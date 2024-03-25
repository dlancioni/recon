import logging
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from sqlite3 import Error
from progress.bar import ShadyBar
from src.baselib import BaseLib
from src.sqllib import SqlLib
from src.fslib import FsLib
from src.dblib import DbLib
from src.utillib import UtilLib
from src.msglib import MsgLib
from src.loglib import LogLib
from src.setuplib import SetupLib
from src.cfglib import ConfigLib
from src.constlib import const
from datetime import datetime, timedelta
import pendulum as pdl

dblib = DbLib()
fslib = FsLib()
sqlib = SqlLib()
msglib = MsgLib()
utillib = UtilLib()
setuplib = SetupLib()
cfglib = ConfigLib()

class EtlLib(BaseLib):

    def __init__(self, id, name):
        self.id = id
        self.name = name
        self.logger = logging.getLogger(__name__)

    def apply_date_pattern(self, file):
        filename = ""
        file_name = setuplib.tag_value(file, "Name").strip()
        mask = setuplib.tag_value(file, "Mask")
        if mask != "":
            mask = cfglib.get_mask(mask)
            start = int(setuplib.tag_value(file, "Start").strip())
            end = int(setuplib.tag_value(file, "End").strip())
            days = int(setuplib.tag_value(file, "Days").strip())
            file_date = file_name[start:end]
            dt = datetime.strptime(file_date, mask)
            if days < 0:
                dt = dt + timedelta(days)
                if dt.weekday() == 5:
                    dt = dt - timedelta(1)
                if dt.weekday() == 6:
                    dt = dt - timedelta(2)
            if days > 0:
                dt = dt + timedelta(days)
                if dt.weekday() == 5:
                    dt = dt + timedelta(1)
                if dt.weekday() == 6:
                    dt = dt + timedelta(2)                    
            file_name = file_name.replace(file_date, dt.strftime(mask))
            filename = file_name
        else:
            filename = file_name
        return filename

    def count(self, file):
        lines = 0
        with open(file, "r") as file:
            lines = len(file.readlines())
        return lines
    
    def persist(self, cn, fields):
        loglib = LogLib("EtlLib", "persist")        
        try:
            field_list = sqlib.get_field_list(fields)
            value_list = sqlib.get_value_list(fields)
            sql = sqlib.get_sql_insert(self.table_name, field_list, value_list)
            rows_affected = dblib.execute(cn, sql)
        except Error as err:
            pass                        
            loglib.log(loglib.INFO, f"Error to manipulate data [{sql}]: {str(err)}")
    
    def get_path(self, datasource):
        if setuplib.tag_value(datasource, "Path", False) == "":
            path = fslib.get_path_file(cfglib.get(3))
        else:
            path = setuplib.tag_value(datasource, "Path", False)
        if fslib.is_dir(path) == False:
            raise Exception(msglib.get("V19", [path]))        
        file = setuplib.tag_value(datasource, "File")
        filename = self.apply_date_pattern(file)
        path = fslib.join(path, filename)
        if fslib.is_file(path) == False:
            raise IOError(msglib.get("V4", [path]))
        return path, filename
    
    def empty_value(self, field_def, field_value=""):
        if str(field_value).strip() in ["", "None"]:
            field_type = setuplib.tag_value(field_def, "Type").strip().lower()
            if field_type in ["integer", "inteiro", "decimal"]:
                field_value = 0
            if field_type in ["text", "texto"]:
                field_value = ""
        return field_value
    
    def default_value(self, field_def, field_value=""):
        default_value = setuplib.tag_value(field_def, "Default Value")
        if str(default_value).strip() != "":
            field_value = str(default_value)
        return field_value
    
    def date_format(self, field_def, field_value=""):
        mask_date = "YYYY-MM-DD"
        mask_time = "HH:mm:SS"
        field_type = setuplib.tag_value(field_def, "Type").lower()
        field_mask = setuplib.tag_value(field_def, "Mask").upper()
        field_mask = field_mask.replace(":MM", ":mm")
        if field_type in const.DATATYPE_DATETIME:
            if field_mask != "":
                if (type(field_value) is datetime.datetime) == True:
                    field_value = pdl.instance(field_value)
                elif (type(field_value) is datetime.time) == True:
                    field_value = str(field_value)
                else:
                    field_value = pdl.from_format(field_value, field_mask)                   
                if field_mask.find("D") == -1 and field_mask.find("M") == -1 and field_mask.find("Y") == -1:
                    mask_date = ""
                if field_mask.find("H") == -1 and field_mask.find("m") == -1 and field_mask.find("S") == -1:
                    mask_time = ""
            field_value = str(field_value.format(f"{mask_date} {mask_time}")).strip()            
        return field_value
    
    def decimal_format(self, field_def, field_value=""):
        field_mask = setuplib.tag_value(field_def, "Mask").strip()
        field_type = setuplib.tag_value(field_def, "Type").lower()
        if field_type in const.DATATYPE_DECIMAL:
            field_value = str(field_value)
            if field_mask == ",":
                field_value = field_value.replace(".", "").replace(",", ".")
        return field_value
    
    def format_data(self, field_def, field_value):
        field_value = self.empty_value(field_def, field_value)
        field_value = self.default_value(field_def, field_value)
        field_value = self.date_format(field_def, field_value)
        field_value = self.decimal_format(field_def, field_value)
        return field_value
    
    def import_delimited(self, cn, datasource):
        row = 0
        field_value = ""
        name = setuplib.tag_value(datasource, "Name")
        path, filename = self.get_path(datasource)
        count = self.count(path)
        fields = self.fields
        start = int(setuplib.tag_value(datasource, "Start"))
        delimiter = setuplib.tag_value(datasource, "Delimiter", False)
        progress_bar = ShadyBar(msglib.set_time(msglib.get("M5", [name, filename])), max=count-1)
        with open(path, "r", encoding='UTF-8') as file:
            for line in file.readlines():
                size = 0
                row += 1
                if (row >= start) and (str(line.strip()) != ""):
                    values = line.split(delimiter) if delimiter != "" else line
                    if len(fields) > len(values):
                        raise Exception(msglib.get("V25", [len(fields), len(values)]))
                    for field in fields:
                        position = int(field[setuplib.tag_name(field, "Position")]) -1
                        field_value = values[position]
                        field["Value"] = self.format_data(field, field_value)
                    self.persist(cn, fields)
                    progress_bar.next()
        progress_bar.finish()
        
    def import_positional(self, cn, datasource):
        row = 0
        field_value = ""
        name = setuplib.tag_value(datasource, "Name")
        path, filename = self.get_path(datasource)
        count = self.count(path)
        fields = self.fields
        start = int(setuplib.tag_value(datasource, "Start"))
        delimiter = setuplib.tag_value(datasource, "Delimiter", False)
        progress_bar = ShadyBar(msglib.set_time(msglib.get("M5", [name, filename])), max=count-1)
        with open(path, "r", encoding='UTF-8') as file:
            for line in file.readlines():
                size = 0
                row += 1
                if (row >= start) and (str(line.strip()) != ""):
                    values = line.split(delimiter) if delimiter != "" else line
                    for field in fields:
                        position = int(field[setuplib.tag_name(field, "Position")]) -1
                        tag_size = setuplib.tag_name(field, "Size", False)
                        if tag_size != "":
                            size = int(field[tag_size])
                            field_value = values[position:position+size]
                        else:    
                            field_value = values[position]
                        field["Value"] = self.format_data(field, field_value)
                    self.persist(cn, fields)
                    progress_bar.next()
        progress_bar.finish()        
        
    def import_db(self, cn, datasource):
        row = 0
        field_value = ""
        fields = self.fields
        name = setuplib.tag_value(datasource, "Name")
        connector = setuplib.tag_value(datasource, "Connector")
        query = setuplib.tag_value(datasource, "Query")
        path = fslib.get_path_task()
        query = fslib.join(path, query)
        if query.find(".sql", 0) >= 0:
            with open(query, 'r') as file:
                query = file.read()
        rows = dblib.get_data(connector, query)
        count = len(rows)
        if count > 0:
            if len(fields) > len(rows[0]):
                raise Exception(msglib.get("V25", [len(fields), len(rows[0])]))
            progress_bar = ShadyBar(msglib.set_time(msglib.get("M5", [name, connector])), max=count)
            for row in rows:
                for field in fields:
                    position = int(field[setuplib.tag_name(field, "Position")]) -1
                    field_value = row[position]
                    field["Value"] = self.format_data(field, field_value)
                self.persist(cn, fields)
                progress_bar.next()
            progress_bar.finish()
        
    def import_excel(self, cn, datasource):
        row = 0
        count = 0        
        field_value = ""
        fields = self.fields
        name = setuplib.tag_value(datasource, "Name")
        start = int(setuplib.tag_value(datasource, "Start"))
        path, filename = self.get_path(datasource)
        msglib.print(msglib.get("M20"))        
        workbook = load_workbook(path)
        sheet_name = setuplib.tag_value(datasource, "Sheet")
        sheet = workbook[sheet_name]
        msglib.print(msglib.get("M19"))
        rows = 65000
        columns = 10
        for row in range(start, rows):
            empty = 0
            for column in range(1, columns):
                if str(sheet.cell(row, column).value).strip() == "None":
                    empty += 1
            if empty == (columns -1):
                break
        count = row
        msg = msglib.get("M5", [name, filename]) + " (" + sheet_name + ")"
        progress_bar = ShadyBar(msglib.set_time(msg), max=count)
        for row in range(1, count):
            if (row >= start):
                for field in fields:
                    column = int(field[setuplib.tag_name(field, "Position")])
                    field_value = sheet.cell(row, column).value
                    field["Value"] = self.format_data(field, field_value)
                self.persist(cn, fields)
                progress_bar.next()                     
        progress_bar.finish()

    def process(self, cn, recon):
        loglib = LogLib("EtlLib", "process")
        try:
            datasources = setuplib.tag_value(recon, "Datasources")
            for datasource in datasources:
                type = setuplib.tag_value(datasource, "Type")
                side = setuplib.tag_value(datasource, "Side")
                self.table_name = f"tb{self.id}{side}"
                self.fields = setuplib.tag_value(datasource, "Fields")
                if type in ["Delimited", "Delimitado"]:
                    self.import_delimited(cn, datasource)
                elif type in ["Positional", "Posicional"]:
                    self.import_positional(cn, datasource)
                elif type in ["Db"]:
                    self.import_db(cn, datasource)
                elif type in ["Excel"]:
                    self.import_excel(cn, datasource)
                else:
                    raise Exception(msglib.get("V20", [type]))
                loglib.log(loglib.INFO, f"File sucessfully imported")

        except Error as err:
            cat = msglib.get("E1")
            msg = f"{cat} {str(err)}"
            loglib.log(loglib.ERROR, msg)
            raise Exception(msg)
        except IOError as err:
            cat = msglib.get("E2")
            msg = f"{cat} {str(err)}"
            loglib.log(loglib.ERROR, msg)
            raise Exception(msg)
        except BaseException as err:
            cat = msglib.get("E3")
            msg = f"{cat} {str(err)}"
            loglib.log(loglib.ERROR, msg)
            raise Exception(msg)